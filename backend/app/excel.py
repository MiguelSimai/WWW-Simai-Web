"""
Genera la planilla que descarga el cliente.

Una fila por expediente, y las columnas las define la plantilla del servicio
(tabla `plantillas_excel`): el cliente de créditos automotrices espera
NumeroOperacion, Patente, una columna por tipo de documento y Observaciones,
y otro cliente esperará otra cosa.

Si un servicio no tiene plantilla, se cae a un juego de columnas genérico: es
mejor entregar algo correcto que un error.
"""

import io
import logging
import unicodedata
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Cuando el servicio no tiene plantilla configurada.
COLUMNAS_GENERICAS: list[dict] = [
    {"titulo": "Solicitud", "origen": "solicitud", "campo": "numero_cliente"},
    {"titulo": "Código", "origen": "solicitud", "campo": "codigo"},
    {"titulo": "Archivos", "origen": "solicitud", "campo": "documentos"},
    {"titulo": "Volumen", "origen": "solicitud", "campo": "unidades"},
    {"titulo": "Costo", "origen": "solicitud", "campo": "costo"},
    {"titulo": "Estado", "origen": "solicitud", "campo": "estado"},
    {"titulo": "Ingresada", "origen": "solicitud", "campo": "creada_en"},
    {"titulo": "Observaciones", "origen": "solicitud", "campo": "resumen"},
]

# Cómo se lee cada estado en la planilla. El cliente no tiene por qué conocer
# el vocabulario interno.
_ESTADO_LEGIBLE = {
    "completada": "OK",
    "revisar": "Revisar",
    "error": "Error",
    "procesando": "En proceso",
}

_ANCHO_MAXIMO = 70
_ANCHO_MINIMO = 12


def _sin_tildes(texto: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn"
    )


def _coincide(nombre_archivo: str, patron: str) -> bool:
    """
    Si el nombre del archivo corresponde al tipo que busca la columna.

    Se compara sin tildes y en mayúsculas porque los nombres vienen tal como
    los dejó quien armó la carpeta: "PAGARE.pdf", "Pagaré.PDF", "pagare (1).pdf"
    son el mismo documento.
    """
    return _sin_tildes(patron).upper() in _sin_tildes(nombre_archivo).upper()


def _valor(columna: dict, solicitud: dict, documentos: list[dict]) -> Any:
    origen = columna.get("origen", "solicitud")
    campo = columna.get("campo", "")

    if origen == "solicitud":
        valor = solicitud.get(campo)
        if campo == "estado":
            return _ESTADO_LEGIBLE.get(str(valor), valor)
        return valor

    if origen == "consolidado":
        datos = solicitud.get("respuesta_ia")
        return _del_json(datos, campo)

    if origen == "documento":
        patron = columna.get("patron") or ""
        hallado = next((d for d in documentos if _coincide(d["archivo"], patron)), None)

        if campo == "presencia":
            return "Sí" if hallado else "No"
        if hallado is None:
            # Documento ausente: celda vacía. Un "No" acá se confundiría con un
            # resultado del análisis.
            return None
        if campo == "estado":
            return _ESTADO_LEGIBLE.get(str(hallado["estado"]), hallado["estado"])
        if campo in ("resultado", "resumen"):
            return hallado.get("resumen") or hallado.get("error")
        return _del_json(hallado.get("respuesta_ia"), campo)

    return None


def _del_json(datos: Any, campo: str) -> Any:
    """
    Saca `campo` del JSON del motor. Acepta rutas con punto ("deudor.rut").

    Una lista se une con "; " para que quepa en una celda: es el caso de
    `observaciones`, que suele venir como arreglo de hallazgos.
    """
    if not isinstance(datos, dict) or not campo:
        return None

    actual: Any = datos
    for parte in campo.split("."):
        if not isinstance(actual, dict):
            return None
        actual = actual.get(parte)
        if actual is None:
            return None

    if isinstance(actual, list):
        return "; ".join(str(x) for x in actual)
    if isinstance(actual, (dict,)):
        return None
    return actual


def _celda(valor: Any) -> Any:
    """Lo que openpyxl puede escribir tal cual."""
    if isinstance(valor, datetime):
        return valor.replace(tzinfo=None)
    if isinstance(valor, (int, float, date, str)) or valor is None:
        return valor
    return str(valor)


def generar(
    filas: list[dict],
    columnas: list[dict],
    titulo_hoja: str = "Operaciones",
) -> bytes:
    """
    Arma el XLSX y lo devuelve en bytes.

    `filas` son dicts de solicitud, cada uno con su lista `documentos_detalle`.
    """
    libro = Workbook()
    hoja = libro.active
    hoja.title = titulo_hoja[:31]  # Excel no admite títulos más largos

    encabezado_fondo = PatternFill("solid", fgColor="F1EAFE")
    encabezado_fuente = Font(bold=True)

    hoja.append([c.get("titulo", "") for c in columnas])
    for celda in hoja[1]:
        celda.fill = encabezado_fondo
        celda.font = encabezado_fuente
        celda.alignment = Alignment(vertical="center")

    for solicitud in filas:
        documentos = solicitud.get("documentos_detalle") or []
        hoja.append([_celda(_valor(c, solicitud, documentos)) for c in columnas])

    # Ancho por contenido, acotado: las observaciones son párrafos y sin tope
    # dejarían una columna de miles de píxeles.
    for i, columna in enumerate(columnas, start=1):
        largos = [len(str(columna.get("titulo", "")))]
        for fila in hoja.iter_rows(min_row=2, min_col=i, max_col=i):
            valor = fila[0].value
            largos.append(len(str(valor)) if valor is not None else 0)

        ancho = min(_ANCHO_MAXIMO, max(_ANCHO_MINIMO, max(largos) + 2))
        hoja.column_dimensions[get_column_letter(i)].width = ancho

        # Las columnas anchas llevan texto ajustado, que es como se lee un
        # párrafo de observaciones en Excel.
        if ancho >= _ANCHO_MAXIMO:
            for fila in hoja.iter_rows(min_row=2, min_col=i, max_col=i):
                fila[0].alignment = Alignment(wrap_text=True, vertical="top")

    # Fija el encabezado al desplazarse: con 100 filas es la diferencia entre
    # leer la planilla y adivinar las columnas.
    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions

    salida = io.BytesIO()
    libro.save(salida)
    return salida.getvalue()
